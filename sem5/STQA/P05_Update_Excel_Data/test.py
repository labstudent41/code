from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from time import sleep
from openpyxl import Workbook

print("Starting Test...")
driver = webdriver.Firefox()
driver.get("http://127.0.0.1:5000/")
sleep(2)

students = [
    ["Ashwin",   20,  "C"],
    ["Azan",     24,  "B"],
    ["Hardik",   24,  "F"],
    ["Laxman",   24,  "C"],
    ["Natwar",   22,  "A"],
    ["Ranjit",   12,  "D"],
    ["Ravi",     20,  "A"],
    ["Russell",  16,  "C"],
    ["Sahil",    19,  "B"],
    ["Saurabh",  13,  "F"],
    ["Suraj",    21,  "D"],
    ["Vaibhav",  17,  "B"],
    ["Vedant",   19,  "B"],
    ["Vikas",    22,  "F"],
    ]

print("Press Enter to continue")
for student in students:
    input("\nStudent: " + str(student))
    input_name = driver.find_element(By.NAME, "name")
    input_age = driver.find_element(By.NAME, "age")
    input_grade = driver.find_element(By.NAME, "grade")

    input_name.send_keys(student[0])
    input_age.send_keys(student[1])
    # input_grade.send_keys(student[2])
    input_grade.submit()
    sleep(1)

    msg = driver.find_element(By.ID, "msg").text
    if msg == "Student record updated successfully!":
        print(student[0], "updated successfully.")
    else:
        input("A wrong input was given, aborting updates...")
        exit()

student_data = []
for row in driver.find_elements(By.XPATH, "//table[@id='dataTable']/tbody/tr"):
    cells = row.find_elements(By.TAG_NAME, "td")
    student_data.append([cells[i].text for i in range(len(cells))])

driver.quit()
wb = Workbook()
ws = wb.active
for i, tag in enumerate(("Name", "Age", "Grade")):
    ws.cell(1, i+1).value = tag

for i, data in enumerate(student_data):
    for j in range(3):
        ws.cell(i+2, j+1).value = data[j]

wb.save("students-grade.xlsx")
print("Data successfully added to sheet.")
